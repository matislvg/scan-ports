"""
scanner_reseau.py — Outil de scan de ports avec rapport Excel colorisé
Auteur  : Matis
Usage   : python scanner_reseau.py -t 192.168.1.0/24 10.0.0.1 -o rapport.xlsx
"""

import argparse
import ipaddress
import logging
import platform
import socket
import subprocess
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration par défaut des ports
# ---------------------------------------------------------------------------

PORTS_PAR_DEFAUT: dict[int, tuple[str, str]] = {
    21:   ("FTP",     "CRITIQUE"),
    22:   ("SSH",     "ÉLEVÉ"),
    23:   ("Telnet",  "CRITIQUE"),
    25:   ("SMTP",    "ÉLEVÉ"),
    53:   ("DNS",     "MODÉRÉ"),
    80:   ("HTTP",    "MODÉRÉ"),
    139:  ("SMB",     "CRITIQUE"),
    443:  ("HTTPS",   "FAIBLE"),
    445:  ("SMB",     "CRITIQUE"),
    3389: ("RDP",     "CRITIQUE"),
    5985: ("WinRM",   "ÉLEVÉ"),
    5986: ("WinRM",   "ÉLEVÉ"),
}

# Ordre de priorité pour la colorisation
CRITICITE_ORDRE = {"CRITIQUE": 0, "ÉLEVÉ": 1, "MODÉRÉ": 2, "FAIBLE": 3}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Résolution des cibles
# ---------------------------------------------------------------------------

def resoudre_cibles(cibles: list[str]) -> list[str]:
    """
    Convertit une liste de cibles (CIDR ou IP seule) en liste d'IPs à scanner.
    Gère correctement les /32 que .hosts() retournerait vides.
    """
    ips: list[str] = []
    for cible in cibles:
        try:
            reseau = ipaddress.ip_network(cible, strict=False)
            hosts = list(reseau.hosts())
            # .hosts() retourne [] pour un /32 ou /128 — on prend l'adresse elle-même
            if hosts:
                ips.extend(str(ip) for ip in hosts)
            else:
                ips.append(str(reseau.network_address))
        except ValueError:
            log.warning("Cible invalide ignorée : %s", cible)
    return ips

# ---------------------------------------------------------------------------
# Découverte d'hôtes
# ---------------------------------------------------------------------------

def ping_icmp(ip: str, timeout: float) -> bool:
    systeme = platform.system().lower()
    if systeme == "windows":
        commande = ["ping", "-n", "1", "-w", str(int(timeout * 1000)), ip]
    else:
        commande = ["ping", "-c", "1", "-W", str(int(timeout)), ip]
    try:
        resultat = subprocess.run(
            commande,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=timeout + 1,
        )
        return resultat.returncode == 0
    except (subprocess.TimeoutExpired, OSError):
        return False


def check_tcp_fallback(ip: str, timeout: float) -> bool:
    """Tente une connexion TCP sur 80 puis 443 comme fallback au PING."""
    for port in (80, 443):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(timeout)
                if s.connect_ex((ip, port)) == 0:
                    return True
        except OSError:
            pass
    return False


def machine_vivante(ip: str, timeout: float) -> tuple[bool, str]:
    if ping_icmp(ip, timeout):
        return True, "PING"
    if check_tcp_fallback(ip, timeout):
        return True, "TCP"
    return False, "-"

# ---------------------------------------------------------------------------
# Résolution hostname
# ---------------------------------------------------------------------------

def resoudre_hostname(ip: str) -> str:
    try:
        return socket.gethostbyaddr(ip)[0]
    except (socket.herror, socket.gaierror):
        return "Inconnu"

# ---------------------------------------------------------------------------
# Scan d'un hôte (exécuté dans un thread)
# ---------------------------------------------------------------------------

def scanner_hote(
    ip: str,
    ports_dict: dict[int, tuple[str, str]],
    timeout: float,
    compteur: list[int],
    total: int,
    lock: threading.Lock,
) -> dict:
    """Scanne un hôte unique. Thread-safe via lock sur le compteur."""
    est_vivante, methode = machine_vivante(ip, timeout)

    infos: dict = {
        "Adresse IP":        ip,
        "Statut":            "UP" if est_vivante else "DOWN",
        "Hostname":          resoudre_hostname(ip) if est_vivante else "-",
        "Méthode détection": methode,
    }

    for port, (service, criticite) in ports_dict.items():
        cle = (port, service, criticite)          # clé interne propre
        if not est_vivante:
            infos[cle] = "-"
            continue
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(timeout)
                infos[cle] = "OUVERT" if s.connect_ex((ip, port)) == 0 else "FERMÉ"
        except (OSError, socket.timeout):
            infos[cle] = "ERREUR"

    with lock:
        compteur[0] += 1
        log.info("[%d/%d] %-16s  Statut=%-4s  Méthode=%s",
                 compteur[0], total, ip, infos["Statut"], methode)

    return infos

# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

COULEURS = {
    "rouge":  PatternFill(start_color="E03C31", end_color="E03C31", fill_type="solid"),
    "orange": PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
    "jaune":  PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "vert":   PatternFill(start_color="90D67B", end_color="90D67B", fill_type="solid"),
    "gris":   PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
    "bleu":   PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),
}

COULEUR_PAR_CRITICITE = {
    "CRITIQUE": ("rouge",  "FFFFFF", True),
    "ÉLEVÉ":    ("orange", "000000", True),
    "MODÉRÉ":   ("jaune",  "000000", False),
    "FAIBLE":   ("bleu",   "000000", False),
}


def generer_excel(
    resultats: list[dict],
    ports_dict: dict[int, tuple[str, str]],
    fichier: str,
) -> None:
    """
    Construit le DataFrame depuis les résultats bruts (clés internes tuple),
    renomme les colonnes en labels lisibles pour l'affichage Excel,
    puis applique le formatage colorisé selon la criticité.
    """
    # --- Renommage des colonnes internes → labels lisibles ---
    rename_map: dict = {}
    for port, (service, criticite) in ports_dict.items():
        cle = (port, service, criticite)
        rename_map[cle] = f"Port {port}\n{service}\n({criticite})"

    df = pd.DataFrame(resultats)

    # Tri par IP (naturel)
    df["_sort"] = df["Adresse IP"].apply(
        lambda ip: int(ipaddress.ip_address(ip))
    )
    df.sort_values("_sort", inplace=True)
    df.drop(columns=["_sort"], inplace=True)

    df.rename(columns=rename_map, inplace=True)

    try:
        with pd.ExcelWriter(fichier, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Rapport")
            ws = writer.sheets["Rapport"]

            # --- En-têtes ---
            header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
            ws.row_dimensions[1].height = 45
            ws.freeze_panes = "A2"

            # --- Largeurs de colonnes ---
            for i, col in enumerate(ws.columns, start=1):
                letter = col[0].column_letter
                ws.column_dimensions[letter].width = 14 if i > 4 else 18

            # --- Colorisation des données ---
            # Reconstruction du mapping label → criticité pour la colorisation
            label_to_criticite: dict[str, str] = {}
            for port, (service, criticite) in ports_dict.items():
                label_to_criticite[f"Port {port}\n{service}\n({criticite})"] = criticite

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue

                    # Colonne Statut
                    header_val = ws.cell(row=1, column=cell.column).value
                    if header_val == "Statut":
                        if val == "UP":
                            cell.fill = COULEURS["vert"]
                            cell.font = Font(bold=True)
                        elif val == "DOWN":
                            cell.fill = COULEURS["gris"]
                            cell.font = Font(color="555555")
                        continue

                    if val == "OUVERT":
                        criticite = label_to_criticite.get(header_val, "FAIBLE")
                        nom_couleur, couleur_texte, bold = COULEUR_PAR_CRITICITE[criticite]
                        cell.fill = COULEURS[nom_couleur]
                        cell.font = Font(bold=bold, color=couleur_texte)

                    elif val == "FERMÉ":
                        cell.fill = COULEURS["vert"]
                        cell.font = Font(color="2D6A2D")

                    elif val == "-":
                        cell.fill = COULEURS["gris"]
                        cell.font = Font(color="888888")

            # --- Alignement centré pour toutes les cellules données ---
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.sheet_view.showGridLines = False

        log.info("Rapport enregistré → %s", fichier)

    except PermissionError:
        log.error("Impossible d'écrire '%s' : fermez le fichier s'il est ouvert.", fichier)
    except Exception as exc:
        log.exception("Erreur inattendue lors de l'export Excel : %s", exc)

# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="scanner_reseau",
        description="Scanner de ports réseau avec rapport Excel colorisé.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python scanner_reseau.py -t 192.168.1.0/24
  python scanner_reseau.py -t 10.0.0.1 10.0.0.2 172.16.0.0/28 -o audit.xlsx
  python scanner_reseau.py -t 192.168.0.0/24 --workers 20 --timeout 0.5
        """,
    )
    parser.add_argument(
        "-t", "--targets", nargs="+", required=True,
        metavar="CIBLE",
        help="Une ou plusieurs cibles : IP seule ou notation CIDR (ex: 192.168.1.0/24)",
    )
    parser.add_argument(
        "-o", "--output", default="rapport_scan.xlsx",
        metavar="FICHIER",
        help="Nom du fichier Excel de sortie (défaut : rapport_scan.xlsx)",
    )
    parser.add_argument(
        "-w", "--workers", type=int, default=20,
        metavar="N",
        help="Nombre de threads parallèles (défaut : 20, max recommandé : 50)",
    )
    parser.add_argument(
        "--timeout", type=float, default=1.0,
        metavar="SEC",
        help="Timeout en secondes par connexion (défaut : 1.0)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Mode verbeux (DEBUG)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Validation
    if args.workers < 1 or args.workers > 200:
        log.error("--workers doit être compris entre 1 et 200.")
        return
    if args.timeout <= 0:
        log.error("--timeout doit être un nombre positif.")
        return

    liste_ip = resoudre_cibles(args.targets)
    if not liste_ip:
        log.error("Aucune IP valide à scanner. Vérifiez vos cibles.")
        return

    total = len(liste_ip)
    workers = min(args.workers, total)
    log.info("Scan de %d machine(s) | %d thread(s) | timeout=%.1fs", total, workers, args.timeout)

    compteur: list[int] = [0]
    lock = threading.Lock()
    resultats: list[dict] = []

    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(
                    scanner_hote,
                    ip, PORTS_PAR_DEFAUT, args.timeout, compteur, total, lock
                ): ip
                for ip in liste_ip
            }
            for future in as_completed(futures):
                try:
                    resultats.append(future.result())
                except Exception as exc:
                    log.error("Erreur sur %s : %s", futures[future], exc)

    except KeyboardInterrupt:
        log.warning("Scan interrompu par l'utilisateur — export des résultats partiels...")

    if not resultats:
        log.warning("Aucun résultat à exporter.")
        return

    generer_excel(resultats, PORTS_PAR_DEFAUT, args.output)


if __name__ == "__main__":
    main()